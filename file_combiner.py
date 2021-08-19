# Weiss Multi-Duty File Scraper

# The purpose of this script is to generate a consolidated spreadsheet containing all item allocations
# pertaining to each container from a Weiss-Rohlig consolidated surcharge invoice. It first
# takes in as input what extract_weiss_files() was able to extract, parses through the desired columns and rows,
# and then outputs the consolidated spreadsheet under /NP-Share/Weiss/ along with a .txt file under /Weiss/Templates
# containing a summary of invoice numbers it was successfully able to find and scrape.

import os
import easygui
import numpy as np

from extract_files import extract_weiss_files
from openpyxl import load_workbook, Workbook
from settings import WEISS_PATH

extracted_files, container_dict, invoice_list, container_list_dict_values = extract_weiss_files()
rows = []
data_rows = []
last_empty_row_list = []


def dialogue_box():
    file_name = easygui.enterbox("What would like you to call your file?")

    return file_name


def run_file_combiner():
    blank_rows = [3, ]

    for i in extracted_files:
        wb_data_only = load_workbook(filename=i, data_only=True)
        sheet_ranges_data_only = wb_data_only['FinalHardCoded']

        last_empty_row = len(list(sheet_ranges_data_only.rows))
        last_empty_row_list.append(last_empty_row)

        cells = sheet_ranges_data_only['A1':'O' + str(last_empty_row)]

        for c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15 in cells:
            rows.append((c1.value, c2.value, c3.value, c4.value, c5.value, c6.value,
                        c7.value, c8.value, c9.value, c10.value, c11.value, c12.value,
                        c13.value, c14.value, c15.value))

    tupled_rows = tuple(rows)
    tupled_data_rows = tuple(data_rows)
    updated_empty_rows = np.cumsum(last_empty_row_list)

    book = Workbook()
    sheet = book.active

    for row in tupled_rows:
        sheet.append(row)

    for amount, updated_last_row in zip(container_list_dict_values, updated_empty_rows):
        # Update Invoice Amount
        sheet.cell(row=updated_last_row - 4, column=8).value = amount

    for updated_last_row in updated_empty_rows:
        # Update CM3 Multiplier Cell
        sheet.cell(row=updated_last_row - 2, column=8).value = \
            '=+H' + str(updated_last_row - 3) + '/F' + str(updated_last_row - 8)
        # Update Minus Duty Cell
        sheet.cell(row=updated_last_row - 3, column=8).value = \
            '=+H' + str(updated_last_row - 4) + '-L' + str(updated_last_row - 7)
        # Update Freight total
        sheet.cell(row=updated_last_row - 3, column=7).value = \
            '=+H' + str(updated_last_row - 4) + '-L' + str(updated_last_row - 7)
        # Update Quantity Checks
        sheet.cell(row=updated_last_row - 6, column=5).value = \
            '=+E' + str(updated_last_row - 8) + '-E' + str(updated_last_row - 7)
        # Update CBM/GW Checks
        sheet.cell(row=updated_last_row - 6, column=6).value = \
            '=+F' + str(updated_last_row - 8) + '-F' + str(updated_last_row - 7)
        # Update Freight Checks
        sheet.cell(row=updated_last_row - 6, column=7).value = \
            '=+G' + str(updated_last_row - 8) + '-G' + str(updated_last_row - 7)
        # Update Factory Invoice Total Checks
        sheet.cell(row=updated_last_row - 6, column=8).value = \
            '=+H' + str(updated_last_row - 8) + '-H' + str(updated_last_row - 7)
        # Update Duty + Tariff Checks
        sheet.cell(row=updated_last_row - 6, column=12).value = \
            '=+L' + str(updated_last_row - 8) + '-L' + str(updated_last_row - 7)
        # Update Freight Amount. Will equal total of the container charges or invoice
        sheet.cell(row=updated_last_row - 7, column=7).value = \
            '=+H' + str(updated_last_row - 4)
        # Update Duty + Tariff Amount. It's 0 because additional charges don't have d/t tacked on
        sheet.cell(row=updated_last_row - 7, column=12).value = 0
        # Update Invoice Check Total
        sheet.cell(row=updated_last_row - 7, column=14).value = \
            '=+H' + str(updated_last_row - 4)

    for updated_last_row, last_row in zip(updated_empty_rows, last_empty_row_list):
        # Update Quantity SUM
        sheet.cell(row=updated_last_row - 8, column=5).value = \
            '=+SUM(E' + str(3 + (updated_last_row - last_row)) + \
            ':E' + str(1 + (updated_last_row - 11)) + ')'
        # Update CBM SUM
        sheet.cell(row=updated_last_row - 8, column=6).value = \
            '=+SUM(F' + str(3 + (updated_last_row - last_row)) + \
            ':F' + str(1 + (updated_last_row - 11)) + ')'
        # Update Freight SUM
        sheet.cell(row=updated_last_row - 8, column=7).value = \
            '=+SUM(G' + str(3 + (updated_last_row - last_row)) + \
            ':G' + str(1 + (updated_last_row - 11)) + ')'
        # Update Invoice SUM
        sheet.cell(row=updated_last_row - 8, column=8).value = \
            '=+SUM(H' + str(3 + (updated_last_row - last_row)) + \
            ':H' + str(1 + (updated_last_row - 11)) + ')'
        # Update Duty + Tariff SUM
        sheet.cell(row=updated_last_row - 8, column=12).value = \
            '=+SUM(L' + str(3 + (updated_last_row - last_row)) + \
            ':L' + str(1 + (updated_last_row - 11)) + ')'
        # Update Freight + Duty SUM
        sheet.cell(row=updated_last_row - 8, column=14).value = \
            '=+SUM(N' + str(3 + (updated_last_row - last_row)) + \
            ':N' + str(1 + (updated_last_row - 11)) + ')'
        # Update Freight SUM
        sheet.cell(row=updated_last_row - 8, column=15).value = \
            '=+SUM(O' + str(3 + (updated_last_row - last_row)) + \
            ':O' + str(1 + (updated_last_row - 11)) + ')'
        for row in range((3 + (updated_last_row - last_row)), (1 + (updated_last_row - 11))):
            # Calculate freight per row
            sheet.cell(row=row, column=7).value = \
                '=F' + str(row) + '*$H$' + str(updated_last_row - 2)
            # Sum duty and freight %'s per row
            sheet.cell(row=row, column=9).value = \
                '=+K' + str(row) + '+J' + str(row)
            # Sum the freight and duty amounts per row
            sheet.cell(row=row, column=12).value = \
                '=H' + str(row) + '*((I' + str(row) + \
                '/100)+0.003464+.00125)'
        for row in range((3 + (updated_last_row - last_row)), (1 + (updated_last_row - 10))):
            if (sheet.cell(row=row, column=1).value is None) and (row not in blank_rows):
                blank_rows.append(row)

    blank_rows_bottom_adjusted = [x - 1 for x in blank_rows[1:]]
    blank_rows_top_adjusted = [x + 1 for x in blank_rows[1:]]
    blank_rows_top_adjusted.insert(0, blank_rows[0])
    blank_rows_top_adjusted.pop()
    blank_rows.pop(0)

    def add_two(list):
        result = []
        for number in list:
            result.append(number + 2)
        result.pop()

        return result

    blank_rows_adj_tupled = list(
        zip(blank_rows[0::1], blank_rows_top_adjusted[0::1], blank_rows_bottom_adjusted[0::1]))

    blank_rows_tupled = zip(blank_rows[0::2], blank_rows[1::2])

    for blank_row, first_of_sum_range, last_of_sum_range in blank_rows_adj_tupled:
        sheet.cell(row=blank_row, column=15).value = '=+SUM(G' + str(first_of_sum_range) + \
            ':G' + str(last_of_sum_range) + ')'

    new_sheet = book.create_sheet(0)
    dialog_answer = dialogue_box()

    output_file_path = os.path.join(WEISS_PATH, dialog_answer + '.xlsx')
    book.save(output_file_path)

    with open(os.path.join(WEISS_PATH, dialog_answer + '.txt'), 'w') as file:
        for inv in invoice_list:
            file.write("%s\n" % inv)

    return output_file_path


run_file_combiner()
